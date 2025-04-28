import requests
from bs4 import BeautifulSoup
import json
import os

# Denna variabel kan du använda för att läsa in data från API'et 
url = "https://realpython.github.io/fake-jobs/"

# Denna variabel kan du använda för att spara ned filen jobs.json i samma mapp som denna fil.
output_path = os.path.join(os.path.dirname(__file__), 'jobs.json')


# INSTRUKTIONER

# Scrapa sidan - https://realpython.github.io/fake-jobs/
# Lägg in det i en lista av dictionaries, där varje dictionary skall innehålla jobtitel, företag, ort, och publiceringsdatum.
# Se till att varje ort är korrekt formaterad med första bokstaven i versal.
# Skriv ned ditt resultat till en JSON-fil kallad job_search_v2.json.

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import os
from collections import defaultdict

# Sätt rätt sökvägar
input_path = r"C:\Users\moab1\AppData\Local\Temp\f17aad81-b9aa-491d-9970-cdb9e8568f75_OOP2_final_home_assignment (4).zip.f75\OOP2_final_home_assignment\exercise_3\sales_data.csv"
output_path = os.path.join(os.path.dirname(__file__), 'sales_output_v2.xlsx')

# Läs in data
sales_data = []
region_totals = defaultdict(float)

with open(input_path, mode='r', encoding='utf-8') as file:
    lines = file.readlines()
    headers = lines[0].strip().split(",")  # Läs rubrikerna
    data_rows = lines[1:]  # Läs raderna med data

    for row in data_rows:
        # Dela upp raden baserat på CSV-format
        row_data = row.strip().split(",")
        record = dict(zip(headers, row_data))

        # Hantera regionbaserad totalförsäljning från `total_price`
        if 'total_price' in record and record['total_price'].isdigit():
            total_price = float(record['total_price'])
        else:
            total_price = 0.0

        region_totals[record['product_list']] += total_price  # Använd `product_list` temporärt

        sales_data.append(record)

# Skapa Excel-dokumentet
wb = Workbook()

# Första arket: Sales Data
ws1 = wb.active
ws1.title = "Sales Data"

# Skriv rubriker med fetstil och storlek 14
for col_index, header in enumerate(headers, start=1):
    cell = ws1.cell(row=1, column=col_index, value=header)
    cell.font = Font(bold=True, size=14)

# Skriv rader med data
for row_index, record in enumerate(sales_data, start=2):
    for col_index, header in enumerate(headers, start=1):
        ws1.cell(row=row_index, column=col_index, value=record.get(header, ""))

# Skapa ett nytt ark för försäljning per region
ws2 = wb.create_sheet(title="Region Sales")

# Skriv rubriker för Region Sales
ws2.append(["Region", "Total Sales"])
for region, total in region_totals.items():
    ws2.append([region, total])

# Formatera rubriker med fetstil och storlek 14
for cell in ws2[1]:
    cell.font = Font(bold=True, size=14)

# Skapa ett diagram för försäljning per region
chart = BarChart()
chart.title = "Total Sales per Region"
chart.x_axis.title = "Region"
chart.y_axis.title = "Total Sales"

data = Reference(ws2, min_col=2, min_row=2, max_row=len(region_totals) + 1)
categories = Reference(ws2, min_col=1, min_row=2, max_row=len(region_totals) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
ws2.add_chart(chart, "E5")  # Placera diagrammet

# Spara filen
wb.save(output_path)
print(f"Excel-fil skapad och sparad till: {output_path}")
